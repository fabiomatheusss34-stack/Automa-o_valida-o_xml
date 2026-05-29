"""
=============================================================
  VALIDADOR DE XML - GRUPO ECOA x T. GLOBO  v2.0
=============================================================
  Requisitos:
    pip install openpyxl pandas xlrd odfpy pyxlsb
=============================================================
"""

import os, glob, re, sys, subprocess
import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import xml.etree.ElementTree as ET
import pandas as pd
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

NS = {"nfe": "http://www.portalfiscal.inf.br/nfe"}

C_HEADER  = "1F3864"
C_GREEN   = "C6EFCE"
C_RED     = "FFC7CE"
C_YELLOW  = "FFEB9C"
C_BLUE    = "DDEEFF"
C_PURPLE  = "EDE7F6"
C_GRAY    = "F2F2F2"
C_ORANGE  = "FFE0CC"

FG_GREEN  = "276221"
FG_RED    = "9C0006"
FG_YELLOW = "7D6608"
FG_BLUE   = "1A4D8F"
FG_PURPLE = "4A1F8C"


# ══════════════════════════════════════════════════════════
#  1. EXTRAÇÃO DO XML
# ══════════════════════════════════════════════════════════

def _txt(el, xpath):
    n = el.find(xpath, NS)
    return n.text.strip() if n is not None and n.text else ""

def _num(v):
    try:
        return float(str(v).replace(",", "."))
    except Exception:
        return 0.0

SEM_EAN_VALS = {"SEM GTIN", "ISENTO", ""}

def extrair_xml(caminho):
    try:
        root = ET.parse(caminho).getroot()
    except ET.ParseError as e:
        return [], f"Erro ao ler XML: {e}"

    nome  = os.path.basename(caminho)
    n_nf  = _txt(root, ".//nfe:nNF")
    emit  = _txt(root, ".//nfe:emit/nfe:xNome")
    data  = _txt(root, ".//nfe:dhEmi")[:10]

    itens = []
    for det in root.findall(".//nfe:det", NS):
        prod = det.find("nfe:prod", NS)
        if prod is None:
            continue

        ean_raw  = _txt(prod, "nfe:cEAN")
        ean_trib = _txt(prod, "nfe:cEANTrib")
        ean_final = (ean_raw  if ean_raw.upper()  not in SEM_EAN_VALS else
                     ean_trib if ean_trib.upper() not in SEM_EAN_VALS else "")

        itens.append({
            "Arquivo XML":    nome,
            "NF":             n_nf,
            "Emitente":       emit,
            "Data Emissão":   data,
            "Nº Item":        det.get("nItem", ""),
            "cProd (SKU)":    _txt(prod, "nfe:cProd"),
            "cEAN (EAN)":     ean_final,
            "cEAN raw":       ean_raw  or "SEM GTIN",
            "cEANTrib":       ean_trib or "SEM GTIN",
            "Sem EAN":        ean_final == "",
            "xProd":          _txt(prod, "nfe:xProd"),
            "NCM":            _txt(prod, "nfe:NCM"),
            "qCom (Qtd)":     _num(_txt(prod, "nfe:qCom")),
            "vUnCom":         _num(_txt(prod, "nfe:vUnCom")),
            "vProd":          _num(_txt(prod, "nfe:vProd")),
        })
    return itens, None


# ══════════════════════════════════════════════════════════
#  2. LEITURA DA PO
# ══════════════════════════════════════════════════════════

def _engine(path):
    ext = os.path.splitext(path)[1].lower()
    if ext in (".xlsx",".xlsm",".xlam",".xltx",".xltm"): return "openpyxl","excel"
    if ext == ".xlsb":   return "pyxlsb","excel"
    if ext in (".xls",".xlt"): return "xlrd","excel"
    if ext in (".ods",".odt"): return "odf","excel"
    if ext in (".csv",".txt"): return None,"csv"
    if ext == ".tsv":          return None,"tsv"
    # Detecta pela assinatura
    try:
        with open(path,"rb") as f: h = f.read(8)
        if h[:4] == b"PK\x03\x04":                          return "openpyxl","excel"
        if h[:8] == b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1":   return "xlrd","excel"
    except Exception: pass
    return "openpyxl","excel"

def _ler_excel(path, eng):
    """
    Lê arquivo Excel, priorizando aba PURCHASE ORDER.
    Cabeçalhos começam na linha 16 e coluna B.
    """
    try:
        xl = pd.ExcelFile(path, engine=eng)

        abas = xl.sheet_names

        aba = next(
            (a for a in abas if "PURCHASE" in a.upper() or a.upper() == "PO"),
            abas[0]
        )

        # ===== LEITURA CORRIGIDA =====
        df = pd.read_excel(
            path,
            sheet_name=aba,
            dtype=str,
            engine=eng,
            header=15  # Linha 16
        )

        # Remove colunas totalmente vazias
        df = df.dropna(axis=1, how="all")

        # Remove espaços dos nomes  
        df.columns = [str(c).strip() for c in df.columns]       
        df.columns = df.columns.str.replace("\n", " ", regex=False)
        df.columns = df.columns.str.replace("\r", " ", regex=False)

        return df, aba

    except Exception:

        # Tenta outros engines
        for e2 in ["openpyxl", "pyxlsb", "xlrd", "odf"]:

            if e2 == eng:
                continue

            try:
                xl = pd.ExcelFile(path, engine=e2)

                abas = xl.sheet_names

                aba = next(
                    (a for a in abas if "PURCHASE" in a.upper() or a.upper() == "PO"),
                    abas[0]
                )

                df = pd.read_excel(
                    path,
                    sheet_name=aba,
                    dtype=str,
                    engine=e2,
                    header=15  # Linha 16
                )

                # Remove colunas vazias
                df = df.dropna(axis=1, how="all")

                # Limpa nomes
                df.columns = [str(c).strip() for c in df.columns]

                return df, aba

            except Exception:
                pass

    raise ValueError("Não foi possível abrir o arquivo da PO.")

def _is_formula(v):
    """Retorna True se a célula contém uma fórmula Excel (começa com =)."""
    return str(v).strip().startswith("=")

def carregar_po(path):
    if not path or not os.path.exists(path):
        return None, None, "Arquivo da PO não encontrado"

    eng, tipo = _engine(path)
    try:
        if tipo == "csv":
            df = pd.read_csv(path, dtype=str, encoding="utf-8-sig")
            aba = "CSV"
        elif tipo == "tsv":
            df = pd.read_csv(path, dtype=str, encoding="utf-8-sig", sep="\t")
            aba = "TSV"
        else:
            df, aba = _ler_excel(path, eng)
    except Exception as e:
        return None, None, str(e)

    df.columns = [str(c).strip() for c in df.columns]

    # ── Localizar colunas fixas ───────────────────────────
    col_sku  = _achar_col(df, ["SKU Number"])
    col_ean  = _achar_col(df, ["Individual Barcode"])
    col_qtd  = _achar_col(df, ["Order quantity","Order Quantity"])

    avisos = []
    if not col_sku:
        col_sku = _achar_col_fuzzy(df, ["sku","cprod","produto"])
        if col_sku: avisos.append(f"SKU via fuzzy → '{col_sku}'")
        else:        avisos.append("❌ Coluna SKU não encontrada")
    if not col_ean:
        col_ean = _achar_col_fuzzy(df, ["barcode","ean","gtin","barras"])
        if col_ean: avisos.append(f"EAN via fuzzy → '{col_ean}'")
        else:        avisos.append("❌ Coluna EAN não encontrada")
    if not col_qtd:
        col_qtd = _achar_col_fuzzy(df, ["order quantity","qtd","quant","qty"])
        if col_qtd: avisos.append(f"Qtd via fuzzy → '{col_qtd}'")
        else:        avisos.append("❌ Coluna Quantidade não encontrada")

    # ── Remover linha de totais/vazias ──────
    linhas_remover = []
    for i, row in df.iterrows():
        sku = str(row[col_sku]).strip() if col_sku else ""
        ean = str(row[col_ean]).strip() if col_ean else ""
        
        is_sku_empty = not sku or sku.lower() == "nan"
        is_ean_empty = not ean or ean.lower() in ("nan", "sem gtin", "isento")
        
        # Considera linha de total/inválida se ambos SKU e EAN estiverem vazios
        if is_sku_empty and is_ean_empty:
            linhas_remover.append(i)
        elif col_qtd and _is_formula(row[col_qtd]):
            linhas_remover.append(i)

    df_limpo = df.drop(index=linhas_remover).copy()
    df_limpo.dropna(how="all", inplace=True)
    df_limpo.reset_index(drop=True, inplace=True)

    if linhas_remover:
        avisos.append(f"Linha(s) de total/inválida(s) removida(s): {len(linhas_remover)} linha(s)")

    # ── Montar df_po padronizado ──────────────────────────
    df_po = pd.DataFrame()
    if col_sku: df_po["SKU_PO"]  = df_limpo[col_sku].astype(str).str.strip()
    if col_ean: df_po["EAN_PO"]  = df_limpo[col_ean].astype(str).str.strip()
    if col_qtd: df_po["QTD_PO"]  = df_limpo[col_qtd].astype(str).str.strip()

    # Guarda df original para aba PO do relatório
    info = {
        "aba": aba,
        "col_sku": col_sku, "col_ean": col_ean, "col_qtd": col_qtd,
        "avisos": avisos,
        "df_original": df_limpo,
        "linhas_formula_removidas": len(linhas_remover),
    }
    return df_po, info, None

def _achar_col(df, nomes):
    for nome in nomes:
        if nome in df.columns: return nome
    return None

def _achar_col_fuzzy(df, keywords):
    for col in df.columns:
        cl = col.lower()
        for kw in keywords:
            if kw in cl: return col
    return None


# ══════════════════════════════════════════════════════════
#  3. LEITURA DA BASE DE RECEBIMENTO
# ══════════════════════════════════════════════════════════

# ── Mapeamento fixo da base de recebimento (recebimento_ge.xlsx) ──────────
# Aba: "rec"  |  Colunas confirmadas pela inspeção do arquivo real
_REC_COLS = {
    "ean":       "Ean",           # col I  — EAN do produto
    "sku":       "SKU",           # col AY — SKU / Cod Produto normalizado
    "cod_prod":  "Cod Produto",   # col J  — código original
    "data":      "Data",          # col D  — data do recebimento
    "nota":      "Nota",          # col N  — número da NF
    "preco":     "Preco",         # col O  — preço unitário
    "quantidade":"Quantidade",    # col P  — quantidade
    "processo":  "Processo",      # col AZ — processo OIKOS
    "evento":    "Evento",        # col G  — tipo de evento
}

def carregar_recebimento(path):
    """
    Lê a base de recebimento (recebimento_ge.xlsx ou similar).
    Prioriza a aba 'rec'. Usa mapeamento fixo de colunas confirmado
    pela inspeção do arquivo real.
    """
    if not path or not os.path.exists(path):
        return None, None, "Arquivo não encontrado"

    eng, tipo = _engine(path)
    df = None
    try:
        if tipo in ("csv","tsv"):
            sep = "\t" if tipo == "tsv" else ","
            df = pd.read_csv(path, dtype=str, encoding="utf-8-sig", sep=sep)
            aba = "CSV"
        else:
            # Priorizar aba "rec"
            try:
                xl  = pd.ExcelFile(path, engine=eng)
                aba = next((a for a in xl.sheet_names if a.lower() == "rec"), xl.sheet_names[0])
                df  = pd.read_excel(path, sheet_name=aba, dtype=str, engine=eng)
            except Exception:
                for e2 in ["openpyxl","pyxlsb","xlrd","odf"]:
                    try:
                        xl  = pd.ExcelFile(path, engine=e2)
                        aba = next((a for a in xl.sheet_names if a.lower() == "rec"), xl.sheet_names[0])
                        df  = pd.read_excel(path, sheet_name=aba, dtype=str, engine=e2)
                        break
                    except Exception:
                        pass
    except Exception as e:
        return None, None, str(e)

    if df is None:
        return None, None, "Não foi possível ler o arquivo de recebimento (formato não suportado ou arquivo corrompido)."

    df.columns = [str(c).strip() for c in df.columns]
    df.dropna(how="all", inplace=True)
    df.reset_index(drop=True, inplace=True)

    # Verifica colunas essenciais — tenta fallback fuzzy se não achar
    mapa = {}
    avisos_rec = []
    for chave, nome_fixo in _REC_COLS.items():
        if nome_fixo in df.columns:
            mapa[chave] = nome_fixo
        else:
            # fuzzy por chave
            fuzzy_map = {
                "ean":        ["ean","gtin","barras","barcode"],
                "sku":        ["sku","cod_prod","cod produto","cod.produto"],
                "cod_prod":   ["cod produto","cod_prod","produto","codigo"],
                "data":       ["data","date","dt"],
                "nota":       ["nota","nf","invoice"],
                "preco":      ["preco","price","valor unit"],
                "quantidade": ["quantidade","qtd","qty","quant"],
                "processo":   ["processo","process"],
                "evento":     ["evento","event","tipo"],
            }
            encontrou = False
            for kw in fuzzy_map.get(chave, []):
                found = _achar_col_fuzzy(df, [kw])
                if found:
                    mapa[chave] = found
                    avisos_rec.append(f"Coluna '{chave}' → encontrada via fuzzy como '{found}'")
                    encontrou = True
                    break
            if not encontrou:
                avisos_rec.append(f"⚠ Coluna '{chave}' (esperado: '{nome_fixo}') não encontrada")

    info = {
        "aba": aba,
        "mapa": mapa,
        "df_original": df,
        "colunas_disponiveis": list(df.columns),
        "avisos": avisos_rec,
    }
    return df, info, None


def analisar_recebimentos(df_rec, info_rec, df_xml):
    """
    Para cada EAN do XML, busca os 2 últimos recebimentos na base.
    Busca SOMENTE por EAN.
    """

    mapa      = info_rec.get("mapa", {})
    col_ean   = mapa.get("ean")
    col_data  = mapa.get("data")
    col_nota  = mapa.get("nota")
    col_qtd   = mapa.get("quantidade")
    col_preco = mapa.get("preco")
    col_proc  = mapa.get("processo")

    if not col_ean:
        return None, "Coluna EAN não identificada na base de recebimento"

    linhas = []

    def _get(row, col, default="—"):
        if row is None:
            return default

        if col and col in df_rec.columns and col in row.index:
            v = str(row[col]).strip()
            return v if v not in ("nan", "None", "") else default

        return default

    def _buscar_historico(ean_xml):
        """
        Busca SOMENTE por EAN.
        Filtra apenas eventos de RECEBIMENTO.
        """

        # EAN inválido
        if not ean_xml or str(ean_xml) in ("", "SEM GTIN", "nan"):
            return pd.DataFrame()

        # Normaliza EAN XML
        ean_xml = re.sub(r"\D", "", str(ean_xml))

        # Normaliza EAN da base
        eans_base = (
            df_rec[col_ean]
            .astype(str)
            .str.replace(r"\D", "", regex=True)
            .str.strip()
        )

        # Busca
        mask = eans_base == ean_xml

        hist = df_rec[mask].copy()

        # Filtra apenas recebimentos
        if "Evento" in hist.columns:

            rec_mask = (
                hist["Evento"]
                .astype(str)
                .str.upper()
                .str.contains("RECEBIMENTO", na=False)
            )

            if rec_mask.any():
                hist = hist[rec_mask]

        # Ordenar por data
        if col_data and col_data in hist.columns:

            try:
                hist["_dt_sort"] = pd.to_datetime(
                    hist[col_data],
                    dayfirst=True,
                    errors="coerce"
                )

                hist = (
                    hist
                    .sort_values("_dt_sort", ascending=False)
                    .drop(columns=["_dt_sort"])
                )

            except Exception:
                pass

        return hist.reset_index(drop=True)

    # Remove duplicados por EAN
    df_xml_uniq = (
        df_xml
        .drop_duplicates(subset=["EAN (XML)"])
        .copy()
    )

    for _, row_xml in df_xml_uniq.iterrows():

        sku_xml = str(row_xml.get("SKU (XML)", "")).strip()

        ean_xml = str(row_xml.get("EAN (XML)", "")).strip()

        # Soma quantidade pelo EAN
        qtd_xml = (
            df_xml[df_xml["EAN (XML)"] == ean_xml]["Qtd (XML)"]
            .sum()
        )

        hist = _buscar_historico(ean_xml)

        # Não encontrado
        if hist.empty:

            linhas.append({
                "SKU (XML)":            sku_xml,
                "EAN (XML)":            ean_xml or "SEM GTIN",
                "Qtd XML atual":        qtd_xml,

                "Último receb. NF":     "Não encontrado na base",
                "Último receb. Data":   "—",
                "Último receb. Qtd":    "—",
                "Último receb. Preço":  "—",
                "Último receb. EAN":    "—",
                "Último receb. Proc.":  "—",

                "Penúlt. receb. NF":    "—",
                "Penúlt. receb. Data":  "—",
                "Penúlt. receb. Qtd":   "—",
                "Penúlt. receb. Preço": "—",
                "Penúlt. receb. EAN":   "—",
                "Penúlt. receb. Proc.": "—",

                "Status Receb.": "⚠ EAN não encontrado na base",
            })

            continue

        ult  = hist.iloc[0]
        penu = hist.iloc[1] if len(hist) > 1 else None

        ean_ult   = _get(ult, col_ean)
        preco_ult = _get(ult, col_preco)

        # Comparação EAN
        if ean_ult == "—" or ean_ult == ean_xml:
            s_ean = "✓ EAN igual ao último receb."
        else:
            s_ean = f"✗ EAN diverge (receb.={ean_ult})"

        linhas.append({

            "SKU (XML)":            sku_xml,
            "EAN (XML)":            ean_xml or "SEM GTIN",
            "Qtd XML atual":        qtd_xml,

            "Último receb. NF":     _get(ult, col_nota),
            "Último receb. Data":   _get(ult, col_data),
            "Último receb. Qtd":    _get(ult, col_qtd),
            "Último receb. Preço":  preco_ult,
            "Último receb. EAN":    ean_ult,
            "Último receb. Proc.":  _get(ult, col_proc),

            "Penúlt. receb. NF":    _get(penu, col_nota),
            "Penúlt. receb. Data":  _get(penu, col_data),
            "Penúlt. receb. Qtd":   _get(penu, col_qtd),
            "Penúlt. receb. Preço": _get(penu, col_preco),
            "Penúlt. receb. EAN":   _get(penu, col_ean),
            "Penúlt. receb. Proc.": _get(penu, col_proc),

            "Status Receb.": s_ean,
        })

    return pd.DataFrame(linhas), None


# ══════════════════════════════════════════════════════════
#  4. VALIDAÇÕES SEPARADAS (PO e XML)
# ══════════════════════════════════════════════════════════

def validar_po(df_po, info_po):
    """
    Valida internamente a PO:
    - SKU duplicado na PO?
    - EAN duplicado na PO?
    - Quantidade zerada?
    - EAN inválido (menos de 8 dígitos)?
    Retorna df com status por linha da PO.
    """
    if df_po is None or df_po.empty:
        return pd.DataFrame()

    df = info_po["df_original"].copy()
    col_sku = info_po["col_sku"]
    col_ean = info_po["col_ean"]
    col_qtd = info_po["col_qtd"]

    status_linhas = []
    skus_vistos = {}
    eans_vistos = {}

    for idx, row in df.iterrows():
        erros = []
        avisos_linha = []

        sku = str(row[col_sku]).strip() if col_sku else ""
        ean = str(row[col_ean]).strip() if col_ean else ""
        qtd_raw = str(row[col_qtd]).strip() if col_qtd else ""

        # SKU duplicado
        if sku and sku != "nan":
            if sku in skus_vistos:
                erros.append(f"SKU duplicado (1ª vez na linha {skus_vistos[sku]+2})")
            else:
                skus_vistos[sku] = idx

        # EAN duplicado
        if ean and ean not in ("nan","SEM GTIN","ISENTO",""):
            if ean in eans_vistos:
                erros.append(f"EAN duplicado (1ª vez na linha {eans_vistos[ean]+2})")
            else:
                eans_vistos[ean] = idx
            # EAN com menos de 8 dígitos
            ean_digits = re.sub(r"\D","",ean)
            if len(ean_digits) < 8:
                erros.append(f"EAN com menos de 8 dígitos ({len(ean_digits)} dígitos)")
        elif ean in ("","nan"):
            avisos_linha.append("Sem EAN")

        # Quantidade zerada ou inválida
        try:
            q = float(qtd_raw.replace(",","."))
            if q <= 0:
                erros.append("Quantidade zerada ou negativa")
        except Exception:
            avisos_linha.append("Quantidade não numérica")

        if erros:
            status = "✗ " + " | ".join(erros)
        elif avisos_linha:
            status = "⚠ " + " | ".join(avisos_linha)
        else:
            status = "✓ OK"

        status_linhas.append({
            "Linha PO":     idx + 2,   # +2: cabeçalho + índice 0-based
            "SKU (PO)":     sku,
            "EAN (PO)":     ean,
            "Qtd (PO)":     qtd_raw,
            "Status PO":    status,
        })

    return pd.DataFrame(status_linhas)


def validar_xml_interno(df_xml):
    """
    Valida internamente o XML:
    - SKU duplicado na mesma NF?
    - EAN duplicado na mesma NF?
    - Quantidade zerada?
    - EAN inválido?
    Retorna df com status por item.
    """
    # Deduplicar por NF + Nº Item
    chave = ["NF","Nº Item"] if "NF" in df_xml.columns else ["Arquivo XML","Nº Item"]
    antes = len(df_xml)
    df = df_xml.drop_duplicates(subset=chave, keep="first").reset_index(drop=True)
    n_dup = antes - len(df)

    status_linhas = []
    # Checagem por NF
    for nf, grupo in df.groupby("NF"):
        # Agrupa itens com mesmo SKU e EAN
        # Usamos copy() e preenchemos NaN para evitar SettingWithCopyWarning
        grupo_limpo = grupo.copy()
        grupo_limpo["cProd (SKU)"] = grupo_limpo["cProd (SKU)"].fillna("")
        grupo_limpo["cEAN (EAN)"] = grupo_limpo["cEAN (EAN)"].fillna("")
        grupo_agg = grupo_limpo.groupby(["cProd (SKU)", "cEAN (EAN)"])
        
        for (sku, ean), itens in grupo_agg:
            erros = []
            avisos_linha = []
            
            sku = str(sku).strip()
            ean = str(ean).strip()
            
            # Soma a quantidade
            qtd_total = itens["qCom (Qtd)"].sum()
            sem_ean = itens["Sem EAN"].iloc[0]
            
            # Combina os números dos itens
            n_items = itens["Nº Item"].astype(str).tolist()
            n_item_str = ", ".join(n_items)
            xprod = itens["xProd"].iloc[0]
            
            if len(n_items) > 1:
                avisos_linha.append(f"Item repetido {len(n_items)} vezes na NF (quantidades somadas)")

            # EAN
            if sem_ean or ean == "":
                # SEM GTIN é um ERRO — produto sem EAN não pode ser aprovado
                erros.append("Produto sem EAN no XML (SEM GTIN) — EAN obrigatório para aprovação")
            else:
                ean_digits = re.sub(r"\D","",ean)
                if len(ean_digits) < 8:
                    erros.append(f"EAN inválido: menos de 8 dígitos ({len(ean_digits)} encontrado(s))")

            # Quantidade
            if qtd_total <= 0:
                erros.append("Quantidade zerada ou negativa")

            if erros:
                status = "✗ " + " | ".join(erros)
            elif avisos_linha:
                status = "⚠ " + " | ".join(avisos_linha)
            else:
                status = "✓ OK"

            status_linhas.append({
                "NF":           nf,
                "Nº Item":      n_item_str,
                "SKU (XML)":    sku,
                "EAN (XML)":    ean if not sem_ean else "SEM GTIN",
                "Qtd (XML)":    qtd_total,
                "xProd":        xprod,
                "Status XML":   status,
            })

    return pd.DataFrame(status_linhas), n_dup


def validar_cruzado(df_xml_status, df_po, df_po_status, df_xml_raw=None, df_rec=None, info_rec=None):
    """
    Cruza XML x PO:
    - EAN do XML == EAN da PO?
    - SKU do XML == SKU da PO?
    - Quantidade do XML == Quantidade da PO?
    Retorna df com resultado por item do XML.
    """
    if df_po is None or df_po.empty:
        df_xml_status = df_xml_status.copy()
        for c in ["EAN PO","SKU PO","Qtd PO","Dif. Qtd",
                  "✓/✗ EAN","✓/✗ SKU","✓/✗ Qtd","Resultado Final"]:
            df_xml_status[c] = "⚠ PO não informada"
        return df_xml_status

    def _norm_ean(v):
        """Normaliza EAN para comparação: remove não-dígitos e zeros à esquerda."""
        return re.sub(r"\D", "", str(v)).lstrip("0") or "0"

    # ── Helpers de preço ──────────────────────────────────────────────────
    def _norm_ean_local(v):
        return re.sub(r"\D", "", str(v)).lstrip("0") or "0"

    def _buscar_preco_xml(ean_xml):
        """Busca vUnCom no df_xml_raw pelo EAN."""
        if df_xml_raw is None or df_xml_raw.empty:
            return None
        col_ean = next((c for c in ["cEAN (EAN)", "EAN (XML)"] if c in df_xml_raw.columns), None)
        if not col_ean or "vUnCom" not in df_xml_raw.columns:
            return None
        ean_norm = _norm_ean_local(ean_xml)
        mask = df_xml_raw[col_ean].astype(str).apply(_norm_ean_local) == ean_norm
        sub  = df_xml_raw[mask]
        if sub.empty:
            return None
        try:
            return float(sub["vUnCom"].iloc[0])
        except Exception:
            return None

    def _buscar_preco_rec(ean_xml):
        """Busca o Preco do último recebimento pelo EAN na base de recebimento."""
        if df_rec is None or df_rec.empty or info_rec is None:
            return None
        mapa     = info_rec.get("mapa", {})
        col_ean  = mapa.get("ean")
        col_prec = mapa.get("preco")
        col_data = mapa.get("data")
        if not col_ean or not col_prec:
            return None
        ean_norm = _norm_ean_local(ean_xml)
        mask = df_rec[col_ean].astype(str).apply(_norm_ean_local) == ean_norm
        hist = df_rec[mask].copy()
        # Filtra eventos de recebimento
        if "Evento" in hist.columns:
            rec_mask = hist["Evento"].astype(str).str.upper().str.contains("RECEBIMENTO", na=False)
            if rec_mask.any():
                hist = hist[rec_mask]
        # Ordena por data (mais recente primeiro)
        if col_data and col_data in hist.columns:
            try:
                hist["_dt"] = pd.to_datetime(hist[col_data], dayfirst=True, errors="coerce")
                hist = hist.sort_values("_dt", ascending=False).drop(columns=["_dt"])
            except Exception:
                pass
        if hist.empty:
            return None
        try:
            return float(str(hist[col_prec].iloc[0]).replace(",", "."))
        except Exception:
            return None

    def _calcular_variacao(preco_xml, preco_rec):
        """Calcula variação % entre preço XML e último recebimento."""
        try:
            px = float(preco_xml)
            pr = float(preco_rec)
            if pr == 0:
                return None, "⚠ Preço receb. = 0"
            var = ((px - pr) / pr) * 100
            sinal = "+" if var > 0 else ""
            var_str = f"{sinal}{var:.2f}%"
            if abs(var) <= 5:
                status = f"✓ OK ({var_str})"
            elif abs(var) <= 15:
                status = f"⚠ Atenção ({var_str})"
            else:
                status = f"✗ Divergência ({var_str})"
            return var_str, status
        except Exception:
            return None, "—"

    linhas = []
    for _, row in df_xml_status.iterrows():
        sku_xml = str(row.get("SKU (XML)","")).strip()
        ean_xml = str(row.get("EAN (XML)","")).strip()
        nf_xml  = str(row.get("NF","")).strip()
        qtd_xml = row.get("Qtd (XML)", 0)
        sem_ean = ean_xml in ("SEM GTIN","","nan")
        ean_xml_norm = _norm_ean(ean_xml) if not sem_ean else ""

        # Busca na PO por SKU (referência principal)
        m_sku = pd.DataFrame()
        if "SKU_PO" in df_po.columns:
            m_sku = df_po[df_po["SKU_PO"].astype(str).str.strip() == sku_xml]

        # Busca por EAN (normalizado)
        m_ean = pd.DataFrame()
        if "EAN_PO" in df_po.columns and not sem_ean:
            eans_po_norm = df_po["EAN_PO"].astype(str).apply(_norm_ean)
            m_ean = df_po[eans_po_norm == ean_xml_norm]

        # ── Status SKU ────────────────────────────────────
        if not m_sku.empty:
            s_sku  = "✓ OK"
            sku_po = m_sku.iloc[0]["SKU_PO"]
        else:
            s_sku  = f"✗ SKU '{sku_xml}' não encontrado na PO"
            sku_po = "—"

        # ── Status EAN ────────────────────────────────────
        if sem_ean:
            s_ean  = "⚠ Produto sem EAN no XML (SEM GTIN)"
            ean_po = "—"
        elif not m_ean.empty:
            s_ean  = "✓ OK"
            ean_po = m_ean.iloc[0]["EAN_PO"]
        elif not m_sku.empty and "EAN_PO" in m_sku.columns:
            ean_po = str(m_sku.iloc[0].get("EAN_PO","")).strip()
            if _norm_ean(ean_po) == ean_xml_norm:
                s_ean = "✓ OK"
            else:
                s_ean = (f"✗ EAN do XML ({ean_xml}) diverge do EAN da PO ({ean_po}) "
                         f"para o SKU '{sku_xml}'")
        else:
            s_ean  = f"✗ EAN '{ean_xml}' do XML não encontrado na PO"
            ean_po = "—"

        # ── Status Quantidade ─────────────────────────────
        ref = m_sku if not m_sku.empty else m_ean
        qtd_po = dif = s_qtd = ""
        if not ref.empty and "QTD_PO" in ref.columns:
            try:
                qtd_po = _num(ref.iloc[0]["QTD_PO"])
                dif    = qtd_xml - qtd_po
                if abs(dif) < 0.01:
                    s_qtd = "✓ OK"
                else:
                    s_qtd = (f"✗ Quantidade diverge: XML={qtd_xml:.0f} | PO={qtd_po:.0f} "
                             f"(diferença de {dif:+.0f})")
            except Exception:
                s_qtd = "⚠ Não foi possível comparar as quantidades"
        else:
            s_qtd = "⚠ Coluna de quantidade não mapeada na PO"

        # ── Resultado Final ───────────────────────────────
        # REGRA: aprovado SOMENTE se EAN, SKU e Quantidade estiverem todos OK (✓)
        # Qualquer ⚠ (sem EAN, quantidade não mapeada) ou ✗ resulta em DIVERGÊNCIA
        ok = (s_ean.startswith("✓")
              and s_sku.startswith("✓")
              and s_qtd.startswith("✓"))
        resultado = "✓ APROVADO" if ok else "✗ DIVERGÊNCIA"

        # ── Comparação de preço por item ─────────────────
        ean_para_preco = ean_xml if not sem_ean else sku_xml
        preco_xml_val  = _buscar_preco_xml(ean_para_preco)
        preco_rec_val  = _buscar_preco_rec(ean_para_preco)

        preco_xml_str = f"{preco_xml_val:.4f}" if preco_xml_val is not None else "—"
        preco_rec_str = f"{preco_rec_val:.4f}" if preco_rec_val is not None else "—"

        if preco_xml_val is not None and preco_rec_val is not None:
            var_str, s_preco = _calcular_variacao(preco_xml_val, preco_rec_val)
        else:
            var_str  = "—"
            s_preco  = "⚠ Sem dados" if preco_xml_val is None and preco_rec_val is None else                        "⚠ Sem preço no XML" if preco_xml_val is None else                        "⚠ Sem histórico de recebimento"

        linhas.append({
            **row.to_dict(),
            "NF":              nf_xml,
            "SKU PO":          sku_po,
            "EAN PO":          ean_po,
            "Qtd PO":          qtd_po if qtd_po != "" else "—",
            "Dif. Qtd":        f"{dif:+.0f}" if isinstance(dif, float) else "—",
            "✓/✗ EAN":         s_ean,
            "✓/✗ SKU":         s_sku,
            "✓/✗ Qtd":         s_qtd,
            "Preço XML":       preco_xml_str,
            "Preço Últ. Receb.": preco_rec_str,
            "Var. Preço %":    var_str,
            "Status Preço":    s_preco,
            "Resultado Final": resultado,
        })

    return pd.DataFrame(linhas)


# ══════════════════════════════════════════════════════════
#  5. GERAÇÃO DO RELATÓRIO
# ══════════════════════════════════════════════════════════

def _bd():
    return Border(left=Side(style="thin"), right=Side(style="thin"),
                  top=Side(style="thin"),  bottom=Side(style="thin"))

def _fill(hex_color):
    return PatternFill("solid", fgColor=hex_color)

def _hdr_cell(cell, valor, wrap=True):
    cell.value = valor
    cell.fill  = _fill(C_HEADER)
    cell.font  = Font(color="FFFFFF", bold=True)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=wrap)
    cell.border = _bd()

def _data_cell(cell, valor, bg=None, bold=False, fg=None, center=True):
    cell.value = valor
    if bg:   cell.fill = _fill(bg)
    cell.font = Font(bold=bold, color=fg or "000000")
    cell.alignment = Alignment(horizontal="center" if center else "left",
                               vertical="center", wrap_text=True)
    cell.border = _bd()

def _cor_status(valor):
    v = str(valor)
    if v.startswith("✓"):   return C_GREEN,  FG_GREEN
    if v.startswith("✗"):   return C_RED,    FG_RED
    if v.startswith("⚠"):   return C_YELLOW, FG_YELLOW
    return None, None

def _tabela(ws, df, colunas, titulo="", cols_status=None):
    """Insere título + cabeçalho + dados numa aba."""
    cols_status = cols_status or []
    linha = 1
    if titulo:
        ws.merge_cells(f"A1:{get_column_letter(len(colunas))}1")
        c = ws["A1"]
        c.value = titulo
        c.fill  = _fill(C_HEADER)
        c.font  = Font(color="FFFFFF", bold=True, size=12)
        c.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 22
        linha = 2

    # Cabeçalho
    for j, col in enumerate(colunas, 1):
        _hdr_cell(ws.cell(row=linha, column=j), col)
    ws.row_dimensions[linha].height = 28
    linha += 1

    # Dados
    df_s = df[[c for c in colunas if c in df.columns]].fillna("").copy()
    for i, (_, row) in enumerate(df_s.iterrows()):
        bg_base = C_GRAY if i % 2 == 1 else None
        for j, col in enumerate(colunas, 1):
            v    = row.get(col, "")
            cell = ws.cell(row=linha, column=j)
            if col in cols_status:
                bg, fg = _cor_status(v)
                _data_cell(cell, v, bg=bg, fg=fg, bold=bool(bg))
            else:
                _data_cell(cell, v, bg=bg_base)
        linha += 1

    # Largura automática
    for j, col in enumerate(colunas, 1):
        vals = [str(df_s.iloc[i].get(col,"")) for i in range(min(len(df_s),80))]
        w    = min(max(len(col), *(len(v) for v in vals)) + 4, 45)
        ws.column_dimensions[get_column_letter(j)].width = w


def gerar_relatorio(df_xml_raw, df_xml_status, df_po_status, df_cruzado,
                    df_recebimento, caminho_saida,
                    arquivos_xml, caminho_po, caminho_rec, nome_processo,
                    info_po, n_dup):

    wb = Workbook()

    # ════════════════════════════════════════════════
    #  ABA 1 — RESUMO
    # ════════════════════════════════════════════════
    ws = wb.active
    ws.title = "RESUMO"
    ws.column_dimensions["A"].width = 30
    ws.column_dimensions["B"].width = 50

    def _meta(row, k, v, bold_v=False, bg_v=None, fg_v=None):
        ws[f"A{row}"] = k
        ws[f"A{row}"].font = Font(bold=True)
        ws[f"B{row}"] = v
        if bg_v: ws[f"B{row}"].fill = _fill(bg_v)
        if fg_v: ws[f"B{row}"].font = Font(bold=bold_v, color=fg_v)
        elif bold_v: ws[f"B{row}"].font = Font(bold=True)

    ws.merge_cells("A1:B1")
    ws["A1"] = "RELATÓRIO DE VALIDAÇÃO XML — GRUPO ECOA"
    ws["A1"].fill = _fill(C_HEADER)
    ws["A1"].font = Font(color="FFFFFF", bold=True, size=14)
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 32

    _meta(2,  "Processo",       nome_processo or "—")
    _meta(3,  "Gerado em",      datetime.now().strftime("%d/%m/%Y %H:%M:%S"))
    _meta(4,  "XMLs processados", ", ".join(os.path.basename(x) for x in arquivos_xml))
    _meta(5,  "PO utilizada",   os.path.basename(caminho_po) if caminho_po else "Não informada")
    _meta(6,  "Base Recebimento", os.path.basename(caminho_rec) if caminho_rec else "Não informada")
    if n_dup: _meta(7, "Itens duplicados removidos", n_dup)

    # Métricas
    ws["A9"] = "MÉTRICAS"
    ws["A9"].font = Font(bold=True, size=12)

    total      = len(df_cruzado)
    aprovados  = (df_cruzado.get("Resultado Final","") == "✓ APROVADO").sum() if not df_cruzado.empty else 0
    diverg     = total - aprovados

    erros_xml  = (df_xml_status["Status XML"].str.startswith("✗")).sum() if not df_xml_status.empty else 0
    avisos_xml = (df_xml_status["Status XML"].str.startswith("⚠")).sum() if not df_xml_status.empty else 0
    erros_po   = (df_po_status["Status PO"].str.startswith("✗")).sum()   if not df_po_status.empty else 0
    avisos_po  = (df_po_status["Status PO"].str.startswith("⚠")).sum()   if not df_po_status.empty else 0

    metricas = [
        ("Total de itens (XML)",    total),
        ("✓ Aprovados (cruzamento)",int(aprovados)),
        ("✗ Divergências cruzamento",int(diverg)),
        ("─── Erros internos no XML",  int(erros_xml)),
        ("⚠ Avisos internos no XML",   int(avisos_xml)),
        ("─── Erros internos na PO",   int(erros_po)),
        ("⚠ Avisos internos na PO",    int(avisos_po)),
        ("NFs processadas",          df_xml_raw["NF"].nunique() if "NF" in df_xml_raw.columns else "—"),
    ]
    for i, (k, v) in enumerate(metricas, start=10):
        ws[f"A{i}"] = k
        ws[f"B{i}"] = v
        ws[f"A{i}"].font = Font(bold=True)
        if "Aprovado" in k and isinstance(v, int) and v > 0:
            ws[f"B{i}"].fill = _fill(C_GREEN)
        elif "Divergência" in k and isinstance(v, int) and v > 0:
            ws[f"B{i}"].fill = _fill(C_RED)
            ws[f"B{i}"].font = Font(bold=True)
        elif "Erro" in k and isinstance(v, int) and v > 0:
            ws[f"B{i}"].fill = _fill(C_ORANGE)

    # Resultado final
    ws["A19"] = "RESULTADO GERAL"
    ws["A19"].font = Font(bold=True, size=13)
    if diverg == 0:
        ws["B19"] = "✓ APROVADO — Nenhuma divergência encontrada"
        ws["B19"].fill = _fill(C_GREEN)
        ws["B19"].font = Font(bold=True, color=FG_GREEN, size=12)
    else:
        ws["B19"] = f"✗ DIVERGÊNCIAS ENCONTRADAS ({diverg} item(ns))"
        ws["B19"].fill = _fill(C_RED)
        ws["B19"].font = Font(bold=True, color=FG_RED, size=12)

    # Avisos da PO
    if info_po and info_po.get("avisos"):
        ws["A21"] = "Avisos da PO"
        ws["A21"].font = Font(bold=True)
        ws["B21"] = " | ".join(info_po["avisos"])
        ws["B21"].fill = _fill(C_YELLOW)

    # ════════════════════════════════════════════════
    #  ABA 2 — DADOS XML (brutos)
    # ════════════════════════════════════════════════
    ws2 = wb.create_sheet("DADOS XML")
    cols_xml = ["Arquivo XML","NF","Emitente","Data Emissão","Nº Item",
                "cProd (SKU)","cEAN (EAN)","cEAN raw","cEANTrib",
                "xProd","NCM","qCom (Qtd)","vUnCom","vProd"]
    _tabela(ws2, df_xml_raw, cols_xml, "Dados extraídos dos XMLs (NF-e)")

    # ════════════════════════════════════════════════
    #  ABA 3 — VALIDAÇÃO XML (interna)
    # ════════════════════════════════════════════════
    ws3 = wb.create_sheet("VALIDAÇÃO XML")
    cols_vxml = ["NF","Nº Item","SKU (XML)","EAN (XML)","Qtd (XML)","xProd","Status XML"]
    _tabela(ws3, df_xml_status, cols_vxml,
            "Validação interna do XML — erros dentro da NF-e",
            cols_status=["Status XML"])

    # ════════════════════════════════════════════════
    #  ABA 4 — DADOS PO (brutos)
    # ════════════════════════════════════════════════
    ws4 = wb.create_sheet("DADOS PO")
    if info_po and "df_original" in info_po:
        df_orig = info_po["df_original"]
        cols_po = list(df_orig.columns)
        _tabela(ws4, df_orig, cols_po, f"Dados da Purchase Order (aba: {info_po.get('aba','')})")
    else:
        ws4["A1"] = "PO não informada"

    # ════════════════════════════════════════════════
    #  ABA 5 — VALIDAÇÃO PO (interna)
    # ════════════════════════════════════════════════
    ws5 = wb.create_sheet("VALIDAÇÃO PO")
    cols_vpo = ["Linha PO","SKU (PO)","EAN (PO)","Qtd (PO)","Status PO"]
    _tabela(ws5, df_po_status, cols_vpo,
            "Validação interna da PO — erros dentro da Purchase Order",
            cols_status=["Status PO"])

    # ════════════════════════════════════════════════
    #  ABA 6 — CRUZAMENTO XML x PO
    # ════════════════════════════════════════════════
    ws6 = wb.create_sheet("XML x PO")
    cols_cruz = [
        "NF", "SKU (XML)", "EAN (XML)", "Qtd (XML)",
        "SKU PO", "EAN PO", "Qtd PO", "Dif. Qtd",
        "✓/✗ EAN", "✓/✗ SKU", "✓/✗ Qtd",
        "Preço XML", "Preço Últ. Receb.", "Var. Preço %", "Status Preço",
        "Resultado Final", "Status XML",
    ]
    _tabela(ws6, df_cruzado, cols_cruz,
            "Cruzamento XML x Purchase Order — divergências entre os dois documentos",
            cols_status=["✓/✗ EAN","✓/✗ SKU","✓/✗ Qtd","Status Preço","Resultado Final","Status XML"])

    # ════════════════════════════════════════════════
    #  ABA 7 — ANÁLISE RECEBIMENTO
    # ════════════════════════════════════════════════
    ws7 = wb.create_sheet("RECEBIMENTO")
    if df_recebimento is not None and not df_recebimento.empty:
        cols_rec = ["SKU (XML)","EAN (XML)","Qtd XML atual",
                    "Último receb. NF","Último receb. Data","Último receb. Qtd","Último receb. EAN",
                    "Penúlt. receb. NF","Penúlt. receb. Data","Penúlt. receb. Qtd","Penúlt. receb. EAN",
                    "Status Receb."]
        _tabela(ws7, df_recebimento, cols_rec,
                "Análise de Recebimento — comparativo com últimos 2 recebimentos",
                cols_status=["Status Receb."])
    else:
        ws7["A1"] = "Base de recebimento não informada ou sem dados correspondentes."
        ws7["A1"].font = Font(italic=True)

    # ════════════════════════════════════════════════
    #  ABA 8 — LOG
    # ════════════════════════════════════════════════
    ws8 = wb.create_sheet("LOG")
    cols_log = ["Data/Hora","Processo","XMLs","Total Itens",
                "Aprovados","Divergências","Erros XML","Erros PO","Resultado"]
    for j, c in enumerate(cols_log, 1):
        _hdr_cell(ws8.cell(row=1, column=j), c)
        ws8.column_dimensions[get_column_letter(j)].width = 20

    ws8.append([
        datetime.now().strftime("%d/%m/%Y %H:%M:%S"),
        nome_processo or "—",
        "; ".join(os.path.basename(x) for x in arquivos_xml),
        total, int(aprovados), int(diverg),
        int(erros_xml), int(erros_po),
        "APROVADO" if diverg == 0 else "DIVERGÊNCIA",
    ])

    wb.save(caminho_saida)


# ══════════════════════════════════════════════════════════
#  6. INTERFACE GRÁFICA
# ══════════════════════════════════════════════════════════

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Validador de XML — Grupo Ecoa x T. Globo  v2.0")
        self.resizable(False, False)
        self.configure(bg="#F0F4F8")
        self._build()
        self._centralizar()

    def _centralizar(self):
        self.update_idletasks()
        w,h  = self.winfo_width(), self.winfo_height()
        sw,sh = self.winfo_screenwidth(), self.winfo_screenheight()
        self.geometry(f"{w}x{h}+{(sw-w)//2}+{(sh-h)//2}")

    # ── UI ───────────────────────────────────────────────
    def _build(self):
        # Cabeçalho
        hdr = tk.Frame(self, bg="#1F3864", padx=20, pady=14)
        hdr.pack(fill="x")
        tk.Label(hdr, text="Validador de XML — Grupo Ecoa x T. Globo",
                 bg="#1F3864", fg="white", font=("Arial",14,"bold")).pack()
        tk.Label(hdr, text="v2.0  |  Validação separada: XML • PO • Cruzamento • Recebimento",
                 bg="#1F3864", fg="#A9C4E8", font=("Arial",9)).pack()

        body = tk.Frame(self, bg="#F0F4F8", padx=20, pady=14)
        body.pack(fill="both")

        self._campo(body,    "Nome do Processo:",      "processo")
        self._arquivo(body,  "Pasta com XMLs:",        "pasta_xml", "dir",
                      dica="Selecione a pasta com os arquivos .xml")
        self._arquivo(body,  "Purchase Order (PO):",   "po", "file",
                      tipos=[
                        ("Todos os formatos","*.xlsx *.xlsm *.xlsb *.xls *.ods *.csv *.tsv *.txt"),
                        ("Excel moderno",    "*.xlsx *.xlsm"),
                        ("Excel binário",    "*.xlsb"),
                        ("Excel legado",     "*.xls"),
                        ("LibreOffice",      "*.ods"),
                        ("CSV/TSV",          "*.csv *.tsv *.txt"),
                        ("Todos",            "*.*"),
                      ],
                      dica="Opcional — arquivo da PO (xlsx, xlsb, xls, ods, csv…)")
        self._arquivo(body, "Base de Recebimento:",   "recebimento", "file",
                      tipos=[
                        ("Todos os formatos","*.xlsx *.xlsm *.xlsb *.xls *.ods *.csv *.tsv"),
                        ("Excel",           "*.xlsx *.xlsm *.xlsb *.xls"),
                        ("CSV/TSV",         "*.csv *.tsv"),
                        ("Todos",           "*.*"),
                      ],
                      dica="Opcional — base de recebimento para análise histórica")
        self._arquivo(body, "Salvar relatório em:",   "saida", "save",
                      tipos=[("Excel","*.xlsx")],
                      dica="Onde salvar o relatório gerado")

        # Log
        tk.Label(body, text="Progresso:", bg="#F0F4F8",
                 font=("Arial",9,"bold")).pack(anchor="w", pady=(12,2))
        self.log_txt = tk.Text(body, height=9, width=64, state="disabled",
                               bg="#1E1E1E", fg="#D4D4D4",
                               font=("Consolas",9), relief="flat", bd=0)
        self.log_txt.pack(fill="x")

        # Botões
        bf = tk.Frame(self, bg="#F0F4F8", pady=12)
        bf.pack()
        self.btn_run = tk.Button(bf, text="▶  Validar XML",
                                 command=self._rodar, bg="#1F3864", fg="white",
                                 relief="flat", font=("Arial",11,"bold"),
                                 padx=24, pady=8, cursor="hand2",
                                 activebackground="#2E4FA0", activeforeground="white")
        self.btn_run.pack(side="left", padx=6)
        tk.Button(bf, text="Limpar", command=self._limpar,
                  bg="#E0E0E0", fg="#333", relief="flat",
                  font=("Arial",10), padx=14, pady=8, cursor="hand2").pack(side="left")

    def _campo(self, parent, label, attr):
        f = tk.Frame(parent, bg="#F0F4F8"); f.pack(fill="x", pady=3)
        tk.Label(f, text=label, bg="#F0F4F8", font=("Arial",9,"bold"),
                 width=22, anchor="w").pack(side="left")
        var = tk.StringVar(); setattr(self, f"var_{attr}", var)
        tk.Entry(f, textvariable=var, width=38, relief="solid", bd=1).pack(side="left", padx=4)

    def _arquivo(self, parent, label, attr, modo, tipos=None, dica=""):
        f = tk.Frame(parent, bg="#F0F4F8"); f.pack(fill="x", pady=3)
        tk.Label(f, text=label, bg="#F0F4F8", font=("Arial",9,"bold"),
                 width=22, anchor="w").pack(side="left")
        var = tk.StringVar(); setattr(self, f"var_{attr}", var)
        tk.Entry(f, textvariable=var, width=32, relief="solid", bd=1).pack(side="left", padx=4)
        def browse():
            if modo == "dir":
                v = filedialog.askdirectory(title=dica)
            elif modo == "save":
                v = filedialog.asksaveasfilename(title=dica, defaultextension=".xlsx",
                                                 filetypes=tipos or [("Excel","*.xlsx")])
            else:
                v = filedialog.askopenfilename(title=dica,
                                               filetypes=tipos or [("Todos","*.*")])
            if v:
                var.set(v)
                if modo == "dir" and not self.var_saida.get():
                    proc = self.var_processo.get().strip().replace(" ","_") or "processo"
                    ts   = datetime.now().strftime("%Y%m%d_%H%M")
                    self.var_saida.set(os.path.join(v, f"Validacao_{proc}_{ts}.xlsx"))
        tk.Button(f, text="📂", command=browse, bg="#E8EEF7",
                  relief="flat", padx=6, cursor="hand2").pack(side="left")

    def _log(self, msg, cor=None):
        self.log_txt.configure(state="normal")
        tags = {"verde":"#4EC9B0","vermelho":"#F48771","amarelo":"#DCDCAA","azul":"#9CDCFE"}
        if cor and cor in tags:
            self.log_txt.tag_configure(cor, foreground=tags[cor])
        self.log_txt.insert("end", msg+"\n", cor)
        self.log_txt.see("end")
        self.log_txt.configure(state="disabled")
        self.update()

    def _limpar(self):
        for a in ["processo","pasta_xml","po","recebimento","saida"]:
            getattr(self, f"var_{a}").set("")
        self.log_txt.configure(state="normal")
        self.log_txt.delete("1.0","end")
        self.log_txt.configure(state="disabled")

    # ── EXECUÇÃO ─────────────────────────────────────────
    def _rodar(self):
        pasta   = self.var_pasta_xml.get().strip()
        po_path = self.var_po.get().strip()
        rec_path= self.var_recebimento.get().strip()
        saida   = self.var_saida.get().strip()
        proc    = self.var_processo.get().strip()

        if not pasta or not os.path.isdir(pasta):
            messagebox.showerror("Erro","Selecione uma pasta válida com os XMLs."); return
        if not saida:
            ts    = datetime.now().strftime("%Y%m%d_%H%M%S")
            saida = os.path.join(pasta, f"Validacao_{ts}.xlsx")
            self.var_saida.set(saida)

        self.btn_run.configure(state="disabled", text="⏳  Processando...")
        self.log_txt.configure(state="normal")
        self.log_txt.delete("1.0","end")
        self.log_txt.configure(state="disabled")

        try:
            # ── 1. Ler XMLs ──────────────────────────────
            self._log("━━━ 1. LENDO XMLs ━━━", "azul")
            xmls = glob.glob(os.path.join(pasta, "*.xml"))
            if not xmls:
                self._log("✗ Nenhum .xml encontrado.", "vermelho"); return

            todos = []
            for xp in xmls:
                itens, err = extrair_xml(xp)
                nome = os.path.basename(xp)
                if err:
                    self._log(f"  ✗ {nome}: {err}", "vermelho")
                else:
                    self._log(f"  ✓ {nome}: {len(itens)} item(ns)")
                    todos.extend(itens)

            if not todos:
                self._log("✗ Nenhum item extraído.", "vermelho"); return

            df_xml_raw = pd.DataFrame(todos)
            self._log(f"  → Total: {len(df_xml_raw)} item(ns) | NFs: {df_xml_raw['NF'].nunique()}")

            # ── 2. Validação interna XML ──────────────────
            self._log("\n━━━ 2. VALIDANDO XML INTERNAMENTE ━━━", "azul")
            df_xml_status, n_dup = validar_xml_interno(df_xml_raw)
            erros_xml  = (df_xml_status["Status XML"].str.startswith("✗")).sum()
            avisos_xml = (df_xml_status["Status XML"].str.startswith("⚠")).sum()
            ok_xml     = len(df_xml_status) - erros_xml - avisos_xml
            self._log(f"  ✓ OK: {ok_xml} | ⚠ Avisos: {avisos_xml} | ✗ Erros: {erros_xml}",
                      "verde" if erros_xml == 0 else "vermelho")
            if n_dup: self._log(f"  ⚠ {n_dup} item(ns) duplicado(s) removido(s)", "amarelo")

            # ── 3. Ler PO ────────────────────────────────
            self._log("\n━━━ 3. LENDO PO ━━━", "azul")
            df_po = df_po_status = info_po = None
            if po_path:
                df_po, info_po, err_po = carregar_po(po_path)
                if err_po:
                    self._log(f"  ✗ {err_po}", "vermelho")
                else:
                    cols_ok = [c for c in ["SKU_PO","EAN_PO","QTD_PO"] if df_po is not None and c in df_po.columns]
                    self._log(f"  ✓ Aba: '{info_po['aba']}' | {len(df_po)} linha(s) | Colunas: {', '.join(cols_ok)}")
                    if info_po.get("avisos"):
                        for av in info_po["avisos"]:
                            self._log(f"  ⚠ {av}", "amarelo")

                    # ── 4. Validação interna PO ───────────
                    self._log("\n━━━ 4. VALIDANDO PO INTERNAMENTE ━━━", "azul")
                    df_po_status = validar_po(df_po, info_po)
                    erros_po  = (df_po_status["Status PO"].str.startswith("✗")).sum()
                    avisos_po = (df_po_status["Status PO"].str.startswith("⚠")).sum()
                    ok_po     = len(df_po_status) - erros_po - avisos_po
                    self._log(f"  ✓ OK: {ok_po} | ⚠ Avisos: {avisos_po} | ✗ Erros: {erros_po}",
                              "verde" if erros_po == 0 else "vermelho")
            else:
                self._log("  ⚠ PO não informada — cruzamento será ignorado", "amarelo")
                df_po_status = pd.DataFrame()

            # ── 5. Cruzamento XML x PO ───────────────────
            self._log("\n━━━ 5. CRUZAMENTO XML x PO ━━━", "azul")
            df_cruzado = validar_cruzado(df_xml_status, df_po, df_po_status, df_xml_raw,
                                                  df_rec_raw if 'df_rec_raw' in dir() else None,
                                                  info_rec   if 'info_rec'   in dir() else None)
            diverg = (df_cruzado["Resultado Final"] == "✗ DIVERGÊNCIA").sum()
            aprov  = (df_cruzado["Resultado Final"] == "✓ APROVADO").sum()
            self._log(f"  ✓ Aprovados: {int(aprov)} | ✗ Divergências: {int(diverg)}",
                      "verde" if diverg == 0 else "vermelho")

            # ── 6. Recebimento ───────────────────────────
            df_rec_analise = None
            df_rec_raw     = None
            info_rec       = None
            self._log("\n━━━ 6. ANÁLISE DE RECEBIMENTO ━━━", "azul")
            if rec_path:
                df_rec_raw, info_rec, err_rec = carregar_recebimento(rec_path)
                if err_rec:
                    self._log(f"  ✗ Erro ao ler base de recebimento: {err_rec}", "vermelho")
                else:
                    aba_rec = info_rec.get('aba', '?')
                    self._log(f"  ✓ Base carregada: {len(df_rec_raw)} registros | Aba: '{aba_rec}'")
                    mapa_rec = info_rec.get('mapa', {})
                    # Mostra quais colunas foram mapeadas
                    for chave, col in mapa_rec.items():
                        self._log(f"    • '{chave}' → coluna '{col}'")
                    # Mostra avisos de colunas não encontradas
                    for av in info_rec.get('avisos', []):
                        cor = "vermelho" if "não encontrada" in av else "amarelo"
                        self._log(f"    {av}", cor)
                    if not mapa_rec.get('ean'):
                        self._log("  ✗ ERRO CRÍTICO: Coluna EAN não encontrada na base de recebimento!"
                                  f" Colunas disponíveis: {list(info_rec.get('colunas_disponiveis',[]))}",
                                  "vermelho")
                    else:
                        df_rec_analise, err_an = analisar_recebimentos(df_rec_raw, info_rec, df_xml_status)
                        if err_an:
                            self._log(f"  ⚠ {err_an}", "amarelo")
                        else:
                            nao_enc = (df_rec_analise["Status Receb."].str.contains("não encontrado", case=False)).sum()
                            div_rec = (df_rec_analise["Status Receb."].str.startswith("✗")).sum()
                            self._log(f"  ✓ {len(df_rec_analise)} produto(s) analisado(s)",
                                      "verde" if div_rec == 0 else "vermelho")
                            if nao_enc:
                                self._log(f"  ⚠ {int(nao_enc)} EAN(s) não encontrado(s) na base", "amarelo")
                            if div_rec:
                                self._log(f"  ✗ {int(div_rec)} divergência(s) de EAN vs recebimento", "vermelho")
            else:
                self._log("  ⚠ Base de recebimento não informada — análise ignorada", "amarelo")

            # ── 7. Gerar relatório ───────────────────────
            self._log("\n━━━ 7. GERANDO RELATÓRIO ━━━", "azul")
            gerar_relatorio(
                df_xml_raw, df_xml_status,
                df_po_status if df_po_status is not None else pd.DataFrame(),
                df_cruzado,
                df_rec_analise,
                saida, xmls, po_path, rec_path, proc, info_po, n_dup
            )
            self._log(f"  ✓ Relatório salvo em:\n  {saida}", "verde")

            # ── Resumo final ─────────────────────────────
            self._log(f"\n{'─'*52}")
            self._log(f"  XMLs lidos:         {len(xmls)}")
            self._log(f"  Itens processados:  {len(df_xml_status)}")
            self._log(f"  ✓ Aprovados:        {int(aprov)}", "verde")
            if diverg > 0:
                self._log(f"  ✗ Divergências:     {int(diverg)}", "vermelho")
            if erros_xml > 0:
                self._log(f"  ✗ Erros no XML:     {int(erros_xml)}", "vermelho")
            self._log(f"{'─'*52}")

            if diverg == 0 and erros_xml == 0:
                self._log("  ✅ RESULTADO FINAL: APROVADO", "verde")
                messagebox.showinfo("✅ Aprovado",
                    f"Validação concluída com sucesso!\n\n"
                    f"✅ APROVADO\n"
                    f"{len(df_xml_status)} item(ns) — nenhuma divergência.")
            else:
                self._log("  ❌ RESULTADO FINAL: DIVERGÊNCIAS ENCONTRADAS", "vermelho")
                resp = messagebox.askquestion("❌ Divergências",
                    f"Validação concluída com divergências.\n\n"
                    f"✗ Divergências no cruzamento: {int(diverg)}\n"
                    f"✗ Erros internos no XML: {int(erros_xml)}\n\n"
                    f"Deseja abrir o relatório agora?")
                if resp == "yes":
                    if sys.platform == "win32":
                        os.startfile(saida)
                    else:
                        subprocess.call(["open", saida])

        except Exception as e:
            import traceback
            self._log(f"\n✗ Erro inesperado: {e}", "vermelho")
            self._log(traceback.format_exc(), "vermelho")
            messagebox.showerror("Erro inesperado", str(e))
        finally:
            self.btn_run.configure(state="normal", text="▶  Validar XML")


if __name__ == "__main__":
    app = App()
    app.mainloop()
